"""
generate_dcf_model.py
=====================
Generates a fully formatted Excel workbook (Polycab_DCF_Model.xlsx) containing
a complete DCF financial model for Polycab India Ltd. (NSE: POLYCAB | BSE: 542652).

Usage:
    python generate_dcf_model.py

Output:
    Polycab_DCF_Model.xlsx  (saved in the current working directory)

Requirements:
    openpyxl >= 3.1.0
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# =============================================================================
# CONSTANTS — All model assumptions are defined here for easy modification
# =============================================================================

# --- Company Info ---
COMPANY_NAME = "Polycab India Ltd."
TICKER = "NSE: POLYCAB | BSE: 542652"
CMP = 8400               # Current Market Price (₹)
SHARES_OUTSTANDING = 150.5  # Millions
VALUATION_DATE = "10-Mar-2026"

# --- CAPM / WACC ---
RISK_FREE_RATE = 0.0670      # 6.70% India 10Y G-Sec
BETA = 0.85
EQUITY_RISK_PREMIUM = 0.0700  # 7.00%
COST_OF_EQUITY = RISK_FREE_RATE + BETA * EQUITY_RISK_PREMIUM   # 12.65%
COST_OF_DEBT_PRETAX = 0.0850   # 8.50%
TAX_RATE = 0.2430              # 24.30%
COST_OF_DEBT_POSTTAX = COST_OF_DEBT_PRETAX * (1 - TAX_RATE)   # ~6.43%
EQUITY_WEIGHT = 0.9950         # 99.50%
DEBT_WEIGHT = 0.0050           # 0.50%
WACC = EQUITY_WEIGHT * COST_OF_EQUITY + DEBT_WEIGHT * COST_OF_DEBT_POSTTAX  # ~12.62%

# --- Growth Assumptions ---
REV_GROWTH_PHASE1 = 0.22   # 22% — Years 1-5
REV_GROWTH_PHASE2 = 0.15   # 15% — Years 6-8
TERMINAL_GROWTH = 0.050    # 5.0%

# --- Margin Assumptions ---
EBITDA_MARGIN_PHASE1 = 0.140   # 14.0% — Years 1-5
EBITDA_MARGIN_PHASE2 = 0.145   # 14.5% — Years 6-8
DA_PCT = 0.0135                # D&A as % of Revenue: 1.35%
CAPEX_PCT = 0.050              # Capex as % of Revenue: 5.0%
NWC_PCT = 0.120                # Change in NWC as % of ΔRevenue: 12.0%

# --- Historical Data ---
HIST_YEARS = ["FY2023", "FY2024", "FY2025"]

HIST_REVENUE    = [14108, 18039, 22408]
HIST_EBITDA     = [1985,  2713,  3168]
HIST_DA         = [209,   245,   298]
HIST_EBIT       = [1776,  2468,  2870]
HIST_INTEREST   = [60,    108,   169]
HIST_PBT        = [1716,  2360,  2701]
HIST_TAX        = [445,   557,   655]
HIST_PAT        = [1271,  1803,  2046]
HIST_CFO        = [1452,  1297,  1809]
HIST_CAPEX      = [850,   970,   1231]
HIST_FCF        = [602,   327,   578]
HIST_EQUITY     = [6605,  8144,  9763]
HIST_DEBT       = [65,    41,    49]
HIST_CASH       = [122,   255,   190]
HIST_ASSETS     = [9227,  11758, 13393]
HIST_CUR_ASSETS = [6740,  8512,  9084]
HIST_CUR_LIAB   = [2558,  3369,  3510]
HIST_NWC        = [4182,  5143,  5574]
HIST_ROE        = [0.192, 0.221, 0.210]
HIST_ROCE       = [0.193, 0.227, 0.224]
HIST_DE         = [0.01,  0.01,  0.01]
HIST_CR         = [2.63,  2.53,  2.59]

# --- Base Year Values (FY2025) ---
BASE_REVENUE   = 22408
BASE_CASH      = 190
BASE_DEBT      = 49

# --- FCFE Model Assumptions ---
NET_INCOME_GROWTH_PHASE1 = 0.25  # 25% — Years 1-5
NET_INCOME_GROWTH_PHASE2 = 0.15  # 15% — Years 6-8
BASE_NET_INCOME = 2046
FCFE_CAPEX = 1100   # ₹ Crore/yr — held constant per management guidance
FCFE_DA  = [340, 390, 445, 500, 560, 620, 680, 740]
FCFE_DWC = [220, 245, 270, 300, 330, 350, 370, 400]

# --- Projection Labels ---
PROJ_YEARS = ["FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E", "FY32E", "FY33E"]
N_YEARS = 8

# =============================================================================
# STYLE HELPERS
# =============================================================================

# Colour palette
COL_HEADER_BG   = "1F3864"   # Dark navy blue
COL_SUBHDR_BG   = "2E75B6"   # Medium blue
COL_INPUT_FONT  = "0000FF"   # Blue — editable inputs
COL_OUTPUT_BG   = "C6EFCE"   # Green — key results
COL_ALT_ROW     = "EBF3FB"   # Light blue — alternating rows
COL_WHITE       = "FFFFFF"
COL_BLACK       = "000000"

# Sheet tab colours
TAB_COLOURS = {
    "Assumptions":        "4472C4",
    "Historical Financials": "70AD47",
    "DCF Model (FCFF)":   "ED7D31",
    "FCFE Model":         "7030A0",
    "Sensitivity Analysis": "FF0000",
}

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _font(bold=False, colour=COL_BLACK, size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thick = Side(style="medium", color="1F3864")
    thin  = Side(style="thin",   color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def _pct_fmt(decimals=1):
    return f'0.{"0"*decimals}%'

def _num_fmt():
    return '#,##0'

def _num_fmt1():
    return '#,##0.0'

def _num_fmt2():
    return '#,##0.00'

# Apply styles to a single cell
def style_cell(cell, value=None, bold=False, font_colour=COL_BLACK, bg=None,
               num_fmt=None, halign="center", valign="center", wrap=False,
               border=True, size=10, italic=False):
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, colour=font_colour, size=size, italic=italic)
    cell.alignment = _align(h=halign, v=valign, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _thin_border()
    if num_fmt:
        cell.number_format = num_fmt

# Write a section header spanning multiple columns
def write_header(ws, row, col_start, col_end, text, bg=COL_HEADER_BG,
                 font_colour=COL_WHITE, size=11):
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = _font(bold=True, colour=font_colour, size=size)
    cell.fill = _fill(bg)
    cell.alignment = _align(h="center", v="center")
    cell.border = _thin_border()
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end
        )

# Write a sub-header
def write_subheader(ws, row, col_start, col_end, text):
    write_header(ws, row, col_start, col_end, text,
                 bg=COL_SUBHDR_BG, font_colour=COL_WHITE, size=10)

# Set column widths
def set_col_widths(ws, widths):
    """widths: list of (col_index, width) tuples"""
    for col_idx, w in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# =============================================================================
# SHEET 1 — ASSUMPTIONS
# =============================================================================

def build_assumptions(ws):
    ws.sheet_properties.tabColor = TAB_COLOURS["Assumptions"]
    ws.freeze_panes = "A3"

    # Column widths
    set_col_widths(ws, [(1, 38), (2, 20), (3, 42)])

    # Title row
    write_header(ws, 1, 1, 3,
                 "POLYCAB INDIA LTD. — DCF Model Assumptions",
                 bg=COL_HEADER_BG, font_colour=COL_WHITE, size=13)

    # Column labels
    for col, lbl in enumerate(["Parameter", "Value", "Notes / Source"], start=1):
        style_cell(ws.cell(row=2, column=col), value=lbl,
                   bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE, halign="center")

    row = 3

    def param_row(label, value, note="", is_input=True, pct=False, fmt=None, bold=False):
        nonlocal row
        c_lbl = ws.cell(row=row, column=1, value=label)
        c_lbl.font = _font(bold=bold, size=10)
        c_lbl.alignment = _align(h="left")
        c_lbl.border = _thin_border()

        c_val = ws.cell(row=row, column=2, value=value)
        fc = COL_INPUT_FONT if is_input else COL_BLACK
        c_val.font = _font(bold=bold, colour=fc, size=10)
        c_val.alignment = _align(h="center")
        c_val.border = _thin_border()
        if fmt:
            c_val.number_format = fmt
        elif pct:
            c_val.number_format = _pct_fmt(2)
        else:
            c_val.number_format = _num_fmt2()

        c_note = ws.cell(row=row, column=3, value=note)
        c_note.font = _font(italic=True, colour="595959", size=9)
        c_note.alignment = _align(h="left", wrap=True)
        c_note.border = _thin_border()
        row += 1

    def section(title):
        nonlocal row
        write_subheader(ws, row, 1, 3, title)
        row += 1

    # ── Company Info ──────────────────────────────────────────────────────────
    section("Company Information")
    param_row("Company Name",          COMPANY_NAME,        "Polycab India Limited",                    fmt="@")
    param_row("Ticker",                TICKER,              "NSE / BSE",                                fmt="@")
    param_row("Current Market Price (CMP)", CMP,            "₹ per share, as on 10-Mar-2026",           fmt='₹#,##0')
    param_row("Shares Outstanding",    SHARES_OUTSTANDING,  "Millions",                                 fmt='#,##0.0')
    param_row("Valuation Date",        VALUATION_DATE,      "",                                         fmt="@")

    # ── WACC ─────────────────────────────────────────────────────────────────
    section("Cost of Capital (CAPM / WACC)")
    param_row("Risk-Free Rate (Rf)",               RISK_FREE_RATE,        "India 10Y G-Sec, Mar 2026",         pct=True)
    param_row("Beta (β)",                          BETA,                  "Blended 2Y/5Y estimate",            fmt='0.00')
    param_row("Equity Risk Premium (ERP)",         EQUITY_RISK_PREMIUM,   "Damodaran India ERP",               pct=True)
    param_row("Cost of Equity (Ke)",               COST_OF_EQUITY,        "= Rf + β × ERP",                    pct=True, is_input=False, bold=True)
    param_row("Cost of Debt (Pre-Tax)",            COST_OF_DEBT_PRETAX,   "Weighted avg borrowing rate",       pct=True)
    param_row("Corporate Tax Rate",                TAX_RATE,              "Effective tax rate FY25",           pct=True)
    param_row("Cost of Debt (Post-Tax)",           COST_OF_DEBT_POSTTAX,  "= Kd × (1 – Tax Rate)",             pct=True, is_input=False)
    param_row("Equity Weight",                     EQUITY_WEIGHT,         "% of total capital",                pct=True)
    param_row("Debt Weight",                       DEBT_WEIGHT,           "% of total capital",                pct=True)
    param_row("WACC",                              WACC,                  "= We×Ke + Wd×Kd(post-tax)",         pct=True, is_input=False, bold=True)

    # ── Growth ────────────────────────────────────────────────────────────────
    section("Revenue Growth Assumptions")
    param_row("Revenue Growth — Phase 1 (Y1–Y5)",  REV_GROWTH_PHASE1,    "FY26E–FY30E; in-line with Project Spring guidance",   pct=True)
    param_row("Revenue Growth — Phase 2 (Y6–Y8)",  REV_GROWTH_PHASE2,    "FY31E–FY33E; gradual normalisation",                  pct=True)
    param_row("Terminal Growth Rate (g)",           TERMINAL_GROWTH,      "Long-run nominal GDP proxy",                          pct=True)

    # ── Margins ───────────────────────────────────────────────────────────────
    section("Margin & Reinvestment Assumptions")
    param_row("EBITDA Margin — Phase 1 (Y1–Y5)",   EBITDA_MARGIN_PHASE1, "FY26E–FY30E; slight improvement from FY25 ~14.1%",   pct=True)
    param_row("EBITDA Margin — Phase 2 (Y6–Y8)",   EBITDA_MARGIN_PHASE2, "FY31E–FY33E; FMEG profitability uplift",              pct=True)
    param_row("D&A as % of Revenue",               DA_PCT,               "Historical average ~1.3–1.4%",                       pct=True)
    param_row("Capex as % of Revenue",             CAPEX_PCT,            "Project Spring ₹1,100 Cr/yr; modelled as ~5% rev",    pct=True)
    param_row("ΔWorking Capital as % of ΔRevenue", NWC_PCT,              "Based on FY23–25 NWC/Revenue trend",                  pct=True)

    # ── FCFE ──────────────────────────────────────────────────────────────────
    section("FCFE Model Assumptions")
    param_row("Base Year Net Income (FY25)",        BASE_NET_INCOME,      "₹ Crore; reported PAT",                               fmt='#,##0')
    param_row("Net Income Growth — Phase 1 (Y1–Y5)", NET_INCOME_GROWTH_PHASE1, "FY26E–FY30E",                                   pct=True)
    param_row("Net Income Growth — Phase 2 (Y6–Y8)", NET_INCOME_GROWTH_PHASE2, "FY31E–FY33E",                                   pct=True)
    param_row("Capex (held constant)",             FCFE_CAPEX,            "₹ Crore/yr per management guidance",                  fmt='#,##0')
    param_row("Net Borrowings",                    0,                     "Net-cash company; assumed nil",                       fmt='#,##0')

    # ── Balance Sheet ─────────────────────────────────────────────────────────
    section("Balance Sheet Inputs (FY2025 Actuals)")
    param_row("Cash & Equivalents",               BASE_CASH,             "₹ Crore; FY25 reported",                              fmt='#,##0')
    param_row("Total Debt",                       BASE_DEBT,             "₹ Crore; FY25 reported",                              fmt='#,##0')

    ws.row_dimensions[1].height = 30

# =============================================================================
# SHEET 2 — HISTORICAL FINANCIALS
# =============================================================================

def build_historical(ws):
    ws.sheet_properties.tabColor = TAB_COLOURS["Historical Financials"]
    ws.freeze_panes = "B3"

    set_col_widths(ws, [(1, 35)] + [(i, 16) for i in range(2, 6)])

    # Title
    write_header(ws, 1, 1, 4,
                 "POLYCAB INDIA LTD. — Historical Financials (₹ Crore)",
                 bg=COL_HEADER_BG, size=13)

    # Year headers
    ws.cell(row=2, column=1).value = "Metric"
    ws.cell(row=2, column=1).font  = _font(bold=True, colour=COL_WHITE)
    ws.cell(row=2, column=1).fill  = _fill(COL_SUBHDR_BG)
    ws.cell(row=2, column=1).alignment = _align(h="center")
    ws.cell(row=2, column=1).border = _thin_border()
    for c, yr in enumerate(HIST_YEARS, start=2):
        style_cell(ws.cell(row=2, column=c), value=yr,
                   bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)

    row = 3

    def data_row(label, values, fmt=_num_fmt(), bold=False, alt=False):
        nonlocal row
        bg = COL_ALT_ROW if alt else None
        c = ws.cell(row=row, column=1, value=label)
        c.font = _font(bold=bold, size=10)
        c.alignment = _align(h="left")
        c.border = _thin_border()
        if bg: c.fill = _fill(bg)
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=row, column=i, value=v)
            style_cell(cell, bold=bold, num_fmt=fmt, bg=bg)
        row += 1

    def section(title):
        nonlocal row
        write_subheader(ws, row, 1, 4, title)
        row += 1

    # Income Statement
    section("Income Statement (₹ Crore)")
    data_row("Revenue",              HIST_REVENUE,  bold=True)
    data_row("EBITDA",               HIST_EBITDA,   alt=True)
    data_row("EBITDA Margin",        [e/r for e,r in zip(HIST_EBITDA, HIST_REVENUE)], fmt=_pct_fmt(1), alt=False)
    data_row("Depreciation & Amortisation (D&A)", HIST_DA, alt=True)
    data_row("EBIT",                 HIST_EBIT)
    data_row("Interest Expense",     HIST_INTEREST,  alt=True)
    data_row("PBT",                  HIST_PBT)
    data_row("Tax",                  HIST_TAX,       alt=True)
    data_row("Net Profit (PAT)",     HIST_PAT,       bold=True)

    # Cash Flow
    section("Cash Flow (₹ Crore)")
    data_row("Cash Flow from Operations (CFO)", HIST_CFO)
    data_row("Capital Expenditure (Capex)",     HIST_CAPEX, alt=True)
    data_row("Free Cash Flow (FCF = CFO–Capex)", HIST_FCF,  bold=True)

    # Balance Sheet
    section("Balance Sheet Highlights (₹ Crore)")
    data_row("Total Equity",         HIST_EQUITY)
    data_row("Total Debt",           HIST_DEBT,      alt=True)
    data_row("Cash & Equivalents",   HIST_CASH)
    data_row("Total Assets",         HIST_ASSETS,    alt=True)
    data_row("Current Assets",       HIST_CUR_ASSETS)
    data_row("Current Liabilities",  HIST_CUR_LIAB,  alt=True)
    data_row("Net Working Capital",  HIST_NWC,       bold=True)

    # Key Ratios
    section("Key Ratios")
    data_row("Return on Equity (ROE)",    HIST_ROE,  fmt=_pct_fmt(1))
    data_row("Return on Capital Emp. (ROCE)", HIST_ROCE, fmt=_pct_fmt(1), alt=True)
    data_row("Debt / Equity (D/E)",       HIST_DE,   fmt='0.00')
    data_row("Current Ratio",             HIST_CR,   fmt='0.00',  alt=True)

    ws.row_dimensions[1].height = 30

# =============================================================================
# SHEET 3 — DCF MODEL (FCFF)
# =============================================================================

def _fcff_projection(wacc=WACC, rev_g1=REV_GROWTH_PHASE1, rev_g2=REV_GROWTH_PHASE2,
                     tg=TERMINAL_GROWTH):
    """Return dict of projected values for the DCF model."""
    revenues, ebitdas, das, ebits, nopats, capexs, dnwcs, fcffs = [], [], [], [], [], [], [], []
    prev_rev = BASE_REVENUE
    for i in range(N_YEARS):
        g = rev_g1 if i < 5 else rev_g2
        rev = prev_rev * (1 + g)
        margin = EBITDA_MARGIN_PHASE1 if i < 5 else EBITDA_MARGIN_PHASE2
        ebitda = rev * margin
        da     = rev * DA_PCT
        ebit   = ebitda - da
        nopat  = ebit * (1 - TAX_RATE)
        capex  = rev * CAPEX_PCT
        dnwc   = (rev - prev_rev) * NWC_PCT
        fcff   = nopat + da - capex - dnwc
        revenues.append(rev); ebitdas.append(ebitda); das.append(da)
        ebits.append(ebit); nopats.append(nopat); capexs.append(capex)
        dnwcs.append(dnwc); fcffs.append(fcff)
        prev_rev = rev

    disc = [(1 / (1 + wacc) ** (i + 1)) for i in range(N_YEARS)]
    pv_fcffs = [f * d for f, d in zip(fcffs, disc)]
    sum_pv   = sum(pv_fcffs)
    tv       = fcffs[-1] * (1 + tg) / (wacc - tg)
    pv_tv    = tv * disc[-1]
    ev       = sum_pv + pv_tv
    eq_val   = ev + BASE_CASH - BASE_DEBT
    iv_share = eq_val / SHARES_OUTSTANDING * 10   # Crore → ₹: × 10^7 / (Mn × 10^6) = × 10

    return dict(
        revenues=revenues, ebitdas=ebitdas, das=das, ebits=ebits,
        nopats=nopats, capexs=capexs, dnwcs=dnwcs, fcffs=fcffs,
        disc=disc, pv_fcffs=pv_fcffs, sum_pv=sum_pv,
        tv=tv, pv_tv=pv_tv, ev=ev, eq_val=eq_val, iv_share=iv_share
    )


def build_dcf(ws):
    ws.sheet_properties.tabColor = TAB_COLOURS["DCF Model (FCFF)"]
    ws.freeze_panes = "B4"

    ncols = N_YEARS + 1   # label col + 8 years
    set_col_widths(ws, [(1, 38)] + [(i, 16) for i in range(2, ncols + 2)])

    p = _fcff_projection()

    # ── Title ─────────────────────────────────────────────────────────────────
    write_header(ws, 1, 1, ncols,
                 "POLYCAB INDIA LTD. — DCF Valuation (FCFF Method)  |  ₹ Crore",
                 bg=COL_HEADER_BG, size=13)

    # ── Model Parameters bar ──────────────────────────────────────────────────
    params_text = (f"WACC: {WACC:.2%}  |  Rev Growth Ph1: {REV_GROWTH_PHASE1:.0%}  |  "
                   f"Rev Growth Ph2: {REV_GROWTH_PHASE2:.0%}  |  Terminal Growth: {TERMINAL_GROWTH:.1%}  |  "
                   f"Tax Rate: {TAX_RATE:.1%}  |  Base Year: FY2025")
    write_header(ws, 2, 1, ncols, params_text,
                 bg="D9E1F2", font_colour=COL_BLACK, size=9)

    # ── Column headers ────────────────────────────────────────────────────────
    style_cell(ws.cell(row=3, column=1), value="Line Item",
               bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    for c, yr in enumerate(PROJ_YEARS, start=2):
        style_cell(ws.cell(row=3, column=c), value=yr,
                   bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)

    row = 4

    def data_row(label, values, fmt=_num_fmt(), bold=False, alt=False,
                 output=False, indent=False):
        nonlocal row
        bg = COL_OUTPUT_BG if output else (COL_ALT_ROW if alt else None)
        prefix = "  " if indent else ""
        c = ws.cell(row=row, column=1, value=prefix + label)
        c.font = _font(bold=bold, size=10)
        c.alignment = _align(h="left")
        c.border = _thin_border()
        if bg: c.fill = _fill(bg)
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=row, column=i, value=round(v, 2) if isinstance(v, float) else v)
            style_cell(cell, bold=bold, num_fmt=fmt, bg=bg)
        row += 1

    def section(title):
        nonlocal row
        write_subheader(ws, row, 1, ncols, title)
        row += 1

    # Revenue
    section("Revenue Projections")
    data_row("Base Year Revenue (FY2025): ₹22,408 Cr", [""] * N_YEARS, fmt="@")
    data_row("Revenue Growth Rate",
             [REV_GROWTH_PHASE1] * 5 + [REV_GROWTH_PHASE2] * 3,
             fmt=_pct_fmt(1), alt=True)
    data_row("Revenue (₹ Cr)", p["revenues"], bold=True)

    # P&L Build-up
    section("P&L Build-up")
    data_row("EBITDA Margin",
             [EBITDA_MARGIN_PHASE1] * 5 + [EBITDA_MARGIN_PHASE2] * 3,
             fmt=_pct_fmt(1), indent=True)
    data_row("EBITDA (₹ Cr)",          p["ebitdas"],         alt=True, indent=True)
    data_row("D&A (₹ Cr)",             p["das"],             indent=True)
    data_row("EBIT (₹ Cr)",            p["ebits"],           bold=True, indent=True, alt=True)
    data_row("Tax on EBIT (@ 24.3%)",  [e*TAX_RATE for e in p["ebits"]], indent=True)
    data_row("NOPAT (₹ Cr)",           p["nopats"],          bold=True)

    # FCFF Build-up
    section("FCFF Build-up")
    data_row("(+) D&A (₹ Cr)",          p["das"],   indent=True, alt=True)
    data_row("(–) Capex (₹ Cr)",        p["capexs"], indent=True)
    data_row("(–) Change in NWC (₹ Cr)", p["dnwcs"], indent=True, alt=True)
    data_row("Free Cash Flow to Firm (FCFF) (₹ Cr)", p["fcffs"],
             bold=True, output=True)

    # Discounting
    section("Discounting (WACC = {:.2%})".format(WACC))
    data_row("Year",              list(range(1, N_YEARS + 1)), fmt="0")
    data_row("Discount Factor",   p["disc"],   fmt="0.0000", alt=True)
    data_row("PV of FCFF (₹ Cr)", p["pv_fcffs"], bold=True)

    # Valuation Summary
    section("Valuation Summary")
    row_sum_pv = row
    style_cell(ws.cell(row=row, column=1), value="Sum of PV of FCFF (₹ Cr)",
               bold=True, halign="left", bg=COL_ALT_ROW)
    style_cell(ws.cell(row=row, column=2), value=round(p["sum_pv"], 0),
               bold=True, num_fmt=_num_fmt(), bg=COL_ALT_ROW)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Terminal Value (₹ Cr)", halign="left")
    style_cell(ws.cell(row=row, column=2), value=round(p["tv"], 0), num_fmt=_num_fmt())
    style_cell(ws.cell(row=row, column=3),
               value=f"= FCFF_Y8 × (1+g) / (WACC–g) = {p['fcffs'][-1]:,.0f} × 1.05 / ({WACC:.4f} – 0.05)",
               halign="left", bold=False, border=False)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="PV of Terminal Value (₹ Cr)",
               bold=True, halign="left", bg=COL_ALT_ROW)
    style_cell(ws.cell(row=row, column=2), value=round(p["pv_tv"], 0),
               bold=True, num_fmt=_num_fmt(), bg=COL_ALT_ROW)
    style_cell(ws.cell(row=row, column=3),
               value=f"Discounted at WACC for {N_YEARS} years",
               halign="left", bold=False, bg=COL_ALT_ROW, border=False)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Enterprise Value (EV) (₹ Cr)",
               bold=True, halign="left", bg=COL_OUTPUT_BG)
    style_cell(ws.cell(row=row, column=2), value=round(p["ev"], 0),
               bold=True, num_fmt=_num_fmt(), bg=COL_OUTPUT_BG)
    row += 1

    style_cell(ws.cell(row=row, column=1),
               value=f"(+) Cash & Equivalents (FY25): ₹{BASE_CASH} Cr", halign="left")
    style_cell(ws.cell(row=row, column=2), value=BASE_CASH, num_fmt=_num_fmt())
    row += 1

    style_cell(ws.cell(row=row, column=1),
               value=f"(–) Total Debt (FY25): ₹{BASE_DEBT} Cr", halign="left")
    style_cell(ws.cell(row=row, column=2), value=BASE_DEBT, num_fmt=_num_fmt())
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Equity Value (₹ Cr)",
               bold=True, halign="left", bg=COL_OUTPUT_BG)
    style_cell(ws.cell(row=row, column=2), value=round(p["eq_val"], 0),
               bold=True, num_fmt=_num_fmt(), bg=COL_OUTPUT_BG)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Shares Outstanding (Mn)",
               halign="left", bg=COL_ALT_ROW)
    style_cell(ws.cell(row=row, column=2), value=SHARES_OUTSTANDING,
               num_fmt="#,##0.0", bg=COL_ALT_ROW)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Intrinsic Value per Share (₹)",
               bold=True, size=12, halign="left", bg=COL_OUTPUT_BG)
    style_cell(ws.cell(row=row, column=2), value=round(p["iv_share"], 0),
               bold=True, size=12, num_fmt='₹#,##0', bg=COL_OUTPUT_BG)
    iv_row = row
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Current Market Price (₹)",
               halign="left")
    style_cell(ws.cell(row=row, column=2), value=CMP, num_fmt='₹#,##0')
    row += 1

    upside = (p["iv_share"] - CMP) / CMP
    style_cell(ws.cell(row=row, column=1), value="Upside / (Downside) %",
               bold=True, halign="left",
               bg=COL_OUTPUT_BG if upside >= 0 else "FFCCCC")
    style_cell(ws.cell(row=row, column=2), value=upside,
               bold=True, num_fmt=_pct_fmt(1),
               bg=COL_OUTPUT_BG if upside >= 0 else "FFCCCC")
    row += 1

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


# =============================================================================
# SHEET 4 — FCFE MODEL
# =============================================================================

def _fcfe_projection(ke=None, ni_g1=None, ni_g2=None, tg=None):
    ke   = ke   if ke   is not None else COST_OF_EQUITY
    ni_g1 = ni_g1 if ni_g1 is not None else NET_INCOME_GROWTH_PHASE1
    ni_g2 = ni_g2 if ni_g2 is not None else NET_INCOME_GROWTH_PHASE2
    tg   = tg   if tg   is not None else TERMINAL_GROWTH

    net_incomes, fcfes = [], []
    ni = BASE_NET_INCOME
    for i in range(N_YEARS):
        g  = ni_g1 if i < 5 else ni_g2
        ni = ni * (1 + g)
        da   = FCFE_DA[i]
        dwc  = FCFE_DWC[i]
        fcfe = ni + da - FCFE_CAPEX - dwc   # Net Borrowings = 0
        net_incomes.append(ni)
        fcfes.append(fcfe)

    disc     = [(1 / (1 + ke) ** (i + 1)) for i in range(N_YEARS)]
    pv_fcfes = [f * d for f, d in zip(fcfes, disc)]
    sum_pv   = sum(pv_fcfes)
    tv       = fcfes[-1] * (1 + tg) / (ke - tg)
    pv_tv    = tv * disc[-1]
    eq_val   = sum_pv + pv_tv
    iv_share = eq_val / SHARES_OUTSTANDING * 10   # Crore → ₹: × 10^7 / (Mn × 10^6) = × 10

    return dict(
        net_incomes=net_incomes, fcfes=fcfes, disc=disc,
        pv_fcfes=pv_fcfes, sum_pv=sum_pv, tv=tv, pv_tv=pv_tv,
        eq_val=eq_val, iv_share=iv_share
    )


def build_fcfe(ws):
    ws.sheet_properties.tabColor = TAB_COLOURS["FCFE Model"]
    ws.freeze_panes = "B4"

    ncols = N_YEARS + 1
    set_col_widths(ws, [(1, 40)] + [(i, 16) for i in range(2, ncols + 2)])

    p = _fcfe_projection()

    # Title
    write_header(ws, 1, 1, ncols,
                 "POLYCAB INDIA LTD. — FCFE Valuation (Free Cash Flow to Equity)  |  ₹ Crore",
                 bg=COL_HEADER_BG, size=13)

    params_text = (f"Cost of Equity (Ke): {COST_OF_EQUITY:.2%}  |  "
                   f"Net Income Growth Ph1: {NET_INCOME_GROWTH_PHASE1:.0%}  |  "
                   f"Net Income Growth Ph2: {NET_INCOME_GROWTH_PHASE2:.0%}  |  "
                   f"Terminal Growth: {TERMINAL_GROWTH:.1%}  |  Capex: ₹{FCFE_CAPEX} Cr/yr (constant)")
    write_header(ws, 2, 1, ncols, params_text,
                 bg="EAD1DC", font_colour=COL_BLACK, size=9)

    style_cell(ws.cell(row=3, column=1), value="Line Item",
               bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    for c, yr in enumerate(PROJ_YEARS, start=2):
        style_cell(ws.cell(row=3, column=c), value=yr,
                   bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)

    row = 4

    def data_row(label, values, fmt=_num_fmt(), bold=False, alt=False, output=False, indent=False):
        nonlocal row
        bg = COL_OUTPUT_BG if output else (COL_ALT_ROW if alt else None)
        prefix = "  " if indent else ""
        c = ws.cell(row=row, column=1, value=prefix + label)
        c.font = _font(bold=bold, size=10)
        c.alignment = _align(h="left")
        c.border = _thin_border()
        if bg: c.fill = _fill(bg)
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=row, column=i, value=round(v, 2) if isinstance(v, float) else v)
            style_cell(cell, bold=bold, num_fmt=fmt, bg=bg)
        row += 1

    def section(title):
        nonlocal row
        write_subheader(ws, row, 1, ncols, title)
        row += 1

    section("Net Income Projections")
    data_row("Base Year Net Income (FY25): ₹2,046 Cr", [""] * N_YEARS, fmt="@")
    data_row("Net Income Growth Rate",
             [NET_INCOME_GROWTH_PHASE1] * 5 + [NET_INCOME_GROWTH_PHASE2] * 3,
             fmt=_pct_fmt(1), alt=True)
    data_row("Net Income (₹ Cr)", p["net_incomes"], bold=True)

    section("FCFE Build-up  [FCFE = Net Income + D&A – Capex – ΔWC + Net Borrowings]")
    data_row("Net Income (₹ Cr)",        p["net_incomes"],        indent=True)
    data_row("(+) Depreciation (₹ Cr)",  FCFE_DA,                 indent=True, alt=True)
    data_row("(–) Capex (₹ Cr)",         [FCFE_CAPEX] * N_YEARS,  indent=True)
    data_row("(–) Change in WC (₹ Cr)",  FCFE_DWC,                indent=True, alt=True)
    data_row("(+) Net Borrowings (₹ Cr)", [0] * N_YEARS,           indent=True)
    data_row("Free Cash Flow to Equity (FCFE) (₹ Cr)", p["fcfes"],
             bold=True, output=True)

    section("Discounting (Ke = {:.2%})".format(COST_OF_EQUITY))
    data_row("Year",               list(range(1, N_YEARS + 1)), fmt="0")
    data_row("Discount Factor",    p["disc"],    fmt="0.0000", alt=True)
    data_row("PV of FCFE (₹ Cr)",  p["pv_fcfes"], bold=True)

    section("Valuation Summary")
    for label, value, is_out, is_alt in [
        ("Sum of PV of FCFE (₹ Cr)",      p["sum_pv"],   False, True),
        ("Terminal Value (₹ Cr)",          p["tv"],       False, False),
        ("PV of Terminal Value (₹ Cr)",    p["pv_tv"],    False, True),
        ("Total Equity Value (₹ Cr)",      p["eq_val"],   True,  False),
        ("Shares Outstanding (Mn)",        SHARES_OUTSTANDING, False, True),
    ]:
        bg = COL_OUTPUT_BG if is_out else (COL_ALT_ROW if is_alt else None)
        style_cell(ws.cell(row=row, column=1), value=label,
                   bold=is_out, halign="left", bg=bg)
        style_cell(ws.cell(row=row, column=2),
                   value=round(value, 0) if isinstance(value, float) else value,
                   bold=is_out, num_fmt=_num_fmt(), bg=bg)
        row += 1

    style_cell(ws.cell(row=row, column=1), value="Intrinsic Value per Share (₹)",
               bold=True, size=12, halign="left", bg=COL_OUTPUT_BG)
    style_cell(ws.cell(row=row, column=2), value=round(p["iv_share"], 0),
               bold=True, size=12, num_fmt='₹#,##0', bg=COL_OUTPUT_BG)
    row += 1

    style_cell(ws.cell(row=row, column=1), value="Current Market Price (₹)",
               halign="left")
    style_cell(ws.cell(row=row, column=2), value=CMP, num_fmt='₹#,##0')
    row += 1

    upside = (p["iv_share"] - CMP) / CMP
    style_cell(ws.cell(row=row, column=1), value="Upside / (Downside) %",
               bold=True, halign="left",
               bg=COL_OUTPUT_BG if upside >= 0 else "FFCCCC")
    style_cell(ws.cell(row=row, column=2), value=upside,
               bold=True, num_fmt=_pct_fmt(1),
               bg=COL_OUTPUT_BG if upside >= 0 else "FFCCCC")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


# =============================================================================
# SHEET 5 — SENSITIVITY ANALYSIS
# =============================================================================

def build_sensitivity(ws):
    ws.sheet_properties.tabColor = TAB_COLOURS["Sensitivity Analysis"]

    set_col_widths(ws, [(1, 22)] + [(i, 16) for i in range(2, 8)])

    write_header(ws, 1, 1, 6,
                 "POLYCAB INDIA LTD. — Sensitivity Analysis  |  Intrinsic Value per Share (₹)",
                 bg=COL_HEADER_BG, size=13)

    # ── Table 1: FCFF IV vs WACC & Terminal Growth ────────────────────────────
    wacc_rows = [0.10, 0.11, 0.12, WACC, 0.13, 0.14, 0.15]
    tg_cols   = [0.03, 0.04, 0.05, 0.06, 0.07]

    row = 2
    write_subheader(ws, row, 1, 6,
                    "Table 1: FCFF Intrinsic Value per Share (₹) — WACC vs Terminal Growth Rate")
    row += 1

    # Header row
    style_cell(ws.cell(row=row, column=1), value="WACC \\ Term. Growth",
               bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    for c, tg in enumerate(tg_cols, start=2):
        style_cell(ws.cell(row=row, column=c), value=tg,
                   bold=True, num_fmt=_pct_fmt(0),
                   bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    row += 1

    for wacc_v in wacc_rows:
        is_base_wacc = abs(wacc_v - WACC) < 0.0001
        style_cell(ws.cell(row=row, column=1), value=wacc_v,
                   bold=is_base_wacc, num_fmt=_pct_fmt(2),
                   bg=("FFF2CC" if is_base_wacc else None))
        for c, tg_v in enumerate(tg_cols, start=2):
            is_base_tg = abs(tg_v - TERMINAL_GROWTH) < 0.0001
            is_base = is_base_wacc and is_base_tg
            p = _fcff_projection(wacc=wacc_v, tg=tg_v)
            bg = COL_OUTPUT_BG if is_base else ("FFF2CC" if (is_base_wacc or is_base_tg) else None)
            style_cell(ws.cell(row=row, column=c), value=round(p["iv_share"], 0),
                       bold=is_base, num_fmt='#,##0',
                       bg=bg)
        row += 1

    # ── Table 2: FCFF IV vs WACC & Revenue Growth Phase 1 ────────────────────
    rev_g_cols = [0.18, 0.20, 0.22, 0.25, 0.28]

    row += 1
    write_subheader(ws, row, 1, 6,
                    "Table 2: FCFF Intrinsic Value per Share (₹) — WACC vs Revenue Growth Phase 1 (Y1–Y5)")
    row += 1

    # Header row
    style_cell(ws.cell(row=row, column=1), value="WACC \\ Rev Growth",
               bold=True, bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    for c, rg in enumerate(rev_g_cols, start=2):
        style_cell(ws.cell(row=row, column=c), value=rg,
                   bold=True, num_fmt=_pct_fmt(0),
                   bg=COL_SUBHDR_BG, font_colour=COL_WHITE)
    row += 1

    for wacc_v in wacc_rows:
        is_base_wacc = abs(wacc_v - WACC) < 0.0001
        style_cell(ws.cell(row=row, column=1), value=wacc_v,
                   bold=is_base_wacc, num_fmt=_pct_fmt(2),
                   bg=("FFF2CC" if is_base_wacc else None))
        for c, rg_v in enumerate(rev_g_cols, start=2):
            is_base_rg = abs(rg_v - REV_GROWTH_PHASE1) < 0.0001
            is_base = is_base_wacc and is_base_rg
            p = _fcff_projection(wacc=wacc_v, rev_g1=rg_v)
            bg = COL_OUTPUT_BG if is_base else ("FFF2CC" if (is_base_wacc or is_base_rg) else None)
            style_cell(ws.cell(row=row, column=c), value=round(p["iv_share"], 0),
                       bold=is_base, num_fmt='#,##0',
                       bg=bg)
        row += 1

    # ── Legend ────────────────────────────────────────────────────────────────
    row += 1
    write_header(ws, row, 1, 6, "Legend", bg="595959", font_colour=COL_WHITE, size=9)
    row += 1
    for colour, meaning in [
        (COL_OUTPUT_BG,  "Base case (WACC = {:.2%}, Term. Growth = {:.0%} / Rev Growth = {:.0%})".format(
            WACC, TERMINAL_GROWTH, REV_GROWTH_PHASE1)),
        ("FFF2CC", "Base WACC row or base assumption column"),
    ]:
        c = ws.cell(row=row, column=1, value=meaning)
        c.fill = _fill(colour)
        c.font = _font(size=9)
        c.alignment = _align(h="left")
        c.border = _thin_border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    ws.row_dimensions[1].height = 30

# =============================================================================
# MAIN — Build and save workbook
# =============================================================================

def main():
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheets in order
    sheets = {
        "Assumptions":           wb.create_sheet("Assumptions"),
        "Historical Financials": wb.create_sheet("Historical Financials"),
        "DCF Model (FCFF)":      wb.create_sheet("DCF Model (FCFF)"),
        "FCFE Model":            wb.create_sheet("FCFE Model"),
        "Sensitivity Analysis":  wb.create_sheet("Sensitivity Analysis"),
    }

    print("Building Sheet 1: Assumptions …")
    build_assumptions(sheets["Assumptions"])

    print("Building Sheet 2: Historical Financials …")
    build_historical(sheets["Historical Financials"])

    print("Building Sheet 3: DCF Model (FCFF) …")
    build_dcf(sheets["DCF Model (FCFF)"])

    print("Building Sheet 4: FCFE Model …")
    build_fcfe(sheets["FCFE Model"])

    print("Building Sheet 5: Sensitivity Analysis …")
    build_sensitivity(sheets["Sensitivity Analysis"])

    output_file = "Polycab_DCF_Model.xlsx"
    wb.save(output_file)
    print(f"\n✅  Successfully generated: {output_file}")
    print(f"    Sheets: {', '.join(wb.sheetnames)}")
    print(f"\n    FCFF Intrinsic Value / Share : ₹{_fcff_projection()['iv_share']:,.0f}")
    print(f"    FCFE Intrinsic Value / Share : ₹{_fcfe_projection()['iv_share']:,.0f}")
    print(f"    Current Market Price (CMP)   : ₹{CMP:,}")


if __name__ == "__main__":
    main()
